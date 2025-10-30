from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import random 

random.seed(55)
colors = ["BBFFFF", "AEEEEE", "96CDCD", "668B8B", "98F5FF", "8EE5EE", "7AC5CD", "53868B", "00F5FF", "00E5EE", "00C5CD", "00868B", "00FFFF", "00EEEE", "00CDCD", "008B8B", "97FFFF", "8DEEEE", "79CDCD", "528B8B", "7FFFD4", "76EEC6", "66CDAA", "458B74", "C1FFC1", "B4EEB4", "9BCD9B", "698B69", "54FF9F", "4EEE94", "43CD80", "2E8B57", "9AFF9A", "90EE90", "7CCD7C", "548B54", "00FF7F", "00EE76", "00CD66", "008B45", "00FF00", "00EE00", "00CD00", "008B00", "7FFF00", "76EE00", "66CD00", "458B00", "C0FF3E", "B3EE3A", "9ACD32", "698B22", "CAFF70", "BCEE68", "A2CD5A", "6E8B3D", "FFF68F", "EEE685", "CDC673", "8B864E", "FFEC8B", "EEDC82", "CDBE70", "8B814C", "FFFFE0", "EEEED1", "CDCDB4", "8B8B7A", "FFFF00", "EEEE00", "CDCD00", "8B8B00", "FFD700", "EEC900", "CDAD00", "8B7500", "FFC125", "EEB422", "CD9B1D", "8B6914", "FFB90F", "EEAD0E", "CD950C", "8B658B", "FFC1C1", "EEB4B4", "CD9B9B", "8B6969", "FF6A6A", "EE6363", "CD5555", "8B3A3A", "FF8247", "EE7942", "CD6839", "8B4726", "FFD39B", "EEC591", "CDAA7D", "8B7355", "FFE7BA", "EED8AE", "CDBA96", "8B7E66", "FFA54F", "EE9A49", "CD853F", "8B5A2B", "FF7F24", "EE7621", "CD661D", "8B4513", "FF3030", "EE2C2C", "CD2626", "8B1A1A", "FF4040", "EE3B3B", "CD3333", "8B2323", "FF8C69", "EE8262", "CD7054", "8B4C39", "FFA07A", "EE9572", "CD8162", "8B5742", "FFA500", "EE9A00", "CD8500", "8B5A00", "FF7F00", "EE7600", "CD6600", "8B4500", "FF7256", "EE6A50", "CD5B45", "8B3E2F", "FF6347", "EE5C42", "CD4F39", "8B3626", "FF4500", "EE4000", "CD3700", "8B2500", "FF0000", "EE0000", "CD0000", "8B0000", "FF1493", "EE1289", "CD1076", "8B0A50", "FF6EB4", "EE6AA7", "CD6090", "8B3A62", "FFB5C5", "EEA9B8", "CD919E", "8B636C", "FFAEB9", "EEA2AD", "CD8C95", "8B5F65", "FF82AB", "EE799F", "CD6889", "8B475D", "FF34B3", "EE30A7", "CD2990", "8B1C62", "FF3E96", "EE3A8C", "CD3278", "8B2252", "FF00FF", "EE00EE", "CD00CD", "8B008B", "FF83FA", "EE7AE9", "CD69C9", "8B4789", "FFBBFF", "EEAEEE", "CD96CD", "8B668B", "E066FF", "D15FEE", "B452CD", "7A378B", "BF3EFF", "B23AEE", "9A32CD", "68228B", "9B30FF", "912CEE", "7D26CD", "551A8B", "AB82FF", "9F79EE", "8968CD", "5D478B", "FFE1FF", "EED2EE", "CDB5CD", "8B7B8B", "1C1C1C", "363636", "4F4F4F", "696969", "828282", "9C9C9C", "B5B5B5", "CFCFCF", "E8E8E8", "A9A9A9", "00008B", "008B8B", "8B008B", "8B0000", "FFFAFA", "F8F8FF", "F5F5F5", "DCDCDC", "FFFAF0", "FDF5E6", "FAF0E6", "FAEBD7", "FFEFD5", "FFEBCD", "FFE4C4", "FFDAB9", "FFDEAD", "FFE4B5", "FFF8DC", "FFFFF0", "FFFACD", "FFF5EE", "F0FFF0", "F5FFFA", "F0FFFF", "F0F8FF", "E6E6FA", "FFF0F5", "FFE4E1", "FFFFFF", "000000", "2F4F4F", "696969", "708090", "778899", "BEBEBE", "D3D3D3", "191970", "000080", "6495ED", "483D8B", "6A5ACD", "7B68EE", "8470FF", "0000CD", "4169E1", "0000FF", "1E90FF", "00BFFF", "87CEEB", "87CEFA", "4682B4", "B0C4DE", "ADD8E6", "B0E0E6", "AFEEEE", "00CED1", "48D1CC", "40E0D0", "00FFFF", "E0FFFF", "5F9EA0", "66CDAA", "7FFFD4", "006400", "556B2F", "8FBC8F", "2E8B57", "3CB371", "20B2AA", "98FB98", "00FF7F", "7CFC00", "00FF00", "7FFF00", "00FA9A", "ADFF2F", "32CD32", "9ACD32", "228B22", "6B8E23", "BDB76B", "EEE8AA", "FAFAD2", "FFFFE0", "FFFF00", "FFD700", "EEDD82", "DAA520", "B8860B", "BC8F8F", "CD5C5C", "8B4513", "A0522D", "CD853F", "DEB887", "F5F5DC", "F5DEB3", "F4A460", "D2B48C", "D2691E", "B22222", "A52A2A", "E9967A", "FA8072", "FFA07A", "FFA500", "FF8C00", "FF7F50", "F08080", "FF6347", "FF4500", "FF0000", "FF69B4", "FF1493", "FFC0CB", "FFB6C1", "DB7093", "B03060", "C71585", "D02090", "FF00FF", "EE82EE", "DDA0DD", "DA70D6", "BA55D3", "9932CC", "9400D3", "8A2BE2", "A020F0", "9370DB", "D8BFD8", "FFFAFA", "EEE9E9", "CDC9C9", "8B8989", "FFF5EE", "EEE5DE", "CDC5BF", "8B8682", "FFEFDB", "EEDFCC", "CDC0B0", "8B8378", "FFE4C4", "EED5B7", "CDB79E", "8B7D6B", "FFDAB9", "EECBAD", "CDAF95", "8B7765", "FFDEAD", "EECFA1", "CDB38B", "8B795E", "FFFACD", "EEE9BF", "CDC9A5", "8B8970", "FFF8DC", "EEE8CD", "CDC8B1", "8B8878", "FFFFF0", "EEEEE0", "CDCDC1", "8B8B83", "F0FFF0", "E0EEE0", "C1CDC1", "838B83", "FFF0F5", "EEE0E5", "CDC1C5", "8B8386", "FFE4E1", "EED5D2", "CDB7B5", "8B7D7B", "F0FFFF", "E0EEEE", "C1CDCD", "838B8B", "836FFF", "7A67EE", "6959CD", "473C8B", "4876FF", "436EEE", "3A5FCD", "27408B", "0000FF", "0000EE", "0000CD", "00008B", "1E90FF", "1C86EE", "1874CD", "104E8B", "63B8FF", "5CACEE", "4F94CD", "36648B", "00BFFF", "00B2EE", "009ACD", "00688B", "87CEFF", "7EC0EE", "6CA6CD", "4A708B", "B0E2FF", "A4D3EE", "8DB6CD", "607B8B", "C6E2FF", "B9D3EE", "9FB6CD", "6C7B8B", "CAE1FF", "BCD2EE", "A2B5CD", "6E7B8B", "BFEFFF", "B2DFEE", "9AC0CD", "68838B", "E0FFFF", "D1EEEE"]
random.shuffle(colors)

# [(color, tid)]
def showMemory(table, filename):
  wb = Workbook()
  sheet = wb.active
  # 边框线条
  thin_border = Side(border_style="thin", color="000000")
  cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
  cell_align = Alignment(horizontal="center")  # 居中对齐
  for i in range(len(table)):  # 行
    for j in range(len(table[i])):  # 列
      cell = sheet.cell(row=i+1, column=j+1, value=table[i][j][0])  # 设置单元格的值
      # print(table[i][j][1])
      cell.fill = PatternFill(start_color=colors[table[i][j][1]], fill_type="solid")  # 设置单元格背景色
      cell.border = cell_border
      cell.alignment = cell_align
  wb.save(f"{filename}.xlsx")

def getAtomShape(smem_shape, elem_width=2):
  box_width_byte = smem_shape[1] * elem_width  # 单位字节
  if box_width_byte % 128 == 0:  # swizzle mode 128B
    return (8, 8)
  elif box_width_byte % 64 == 0: # swizzle mode 64B
    return (8, 4)
  elif box_width_byte % 32 == 0: # swizzle mode 32B
    return (8, 2)
  else:
    return (8, 1)

def swizzle(row, col, atom, mode="our"):
  def xor2x2(i, j):
    return (i + j) % 2

  def xor4x4(i, j):
    i0 = i % 2
    j0 = j % 2
    i1 = i // 2
    j1 = j // 2
    return 2 * xor2x2(i1, j1) + xor2x2(i0, j0)

  def xor8x8(i, j):
    i0 = i % 2
    j0 = j % 2
    i1 = i // 2
    j1 = j // 2
    return 2 * xor4x4(i1, j1) + xor2x2(i0, j0)
  
  if mode == "our":
    return col ^ (row % atom[1])
  elif mode == "tilelang":
    if atom[1] == 8:
      return xor8x8(row, col)
    elif atom[1] == 4:
      return xor4x4(row, col)
    elif atom[1] == 2:
      return xor2x2(row, col)
  elif mode == "our_other":
    r = 8 // atom[1]
    return col ^ ((row // r) % atom[0])
  return col

def getSwizzlePhysicsShape(smem_shape, atom, elem_width=2):
  box_width_16B = smem_shape[1] * elem_width // 16  # 16B为单位
  s = box_width_16B // atom[1]
  c = smem_shape[0] // atom[0]
  atom_size = atom[0] * atom[1] * 16 // elem_width  # atom 原始元素个数
  return (s, c, atom_size)

def originSmem(smem_shape, elem_width=2):
  # 原始共享内存布局
  vec_per_elem = 16 // elem_width  # 一个向量化访存的宽度含有几个基础元素的宽度
  table = []
  for i in range(smem_shape[0]):
    table_row = []
    for j in range(smem_shape[1]):
      # 线性化索引
      idx = i * smem_shape[1] + j

      if j % vec_per_elem == 0:
        table_row.append((idx // vec_per_elem, j // vec_per_elem))  # 设置同列颜色相同
    table.append(table_row)
  showMemory(table, "originSmem")

def swizzleSmem(smem_shape, elem_width=2, mode="our"):
  # swizzle后的共享内存布局
  vec_per_elem = 16 // elem_width  # 一个向量化访存的宽度含有几个基础元素的宽度
  atom = getAtomShape(smem_shape, elem_width)
  phy_shape = getSwizzlePhysicsShape(smem_shape, atom, elem_width)
  print(atom, phy_shape)
  table = []
  for i in range(smem_shape[0]):
    table_row = []
    for j in range(smem_shape[1]):
      # 转换为向量化后的列索引
      vec_j = j // vec_per_elem
      # 该元素在向量内的索引
      vec_idx = vec_j % vec_per_elem
      # atomic block的索引
      atom_i = i // atom[0]
      atom_j = vec_j // atom[1]
      # atomic block内的坐标
      atom_inner_i = i % atom[0]
      atom_inner_j = vec_j % atom[1]
      # atom_inner_j 进行swizzle
      atom_inner_j = swizzle(atom_inner_i, atom_inner_j, atom, mode)
      # atom 内一维地址（粒度为elem）
      atom_idx = (atom_inner_i * atom[1] + atom_inner_j) * vec_per_elem + vec_idx
      # 物理地址计算
      idx = atom_j * phy_shape[1] * phy_shape[2] + atom_i * phy_shape[2] + atom_idx

      # 转换回原始逻辑shape的二维坐标
      new_i = idx // smem_shape[1]
      new_j = idx % smem_shape[1]
      if j % vec_per_elem == 0:
        table_row.append((idx // vec_per_elem, new_j // vec_per_elem))  # 设置同列颜色相同
    table.append(table_row)
  showMemory(table, "swizzleSmem")
  

if __name__ == '__main__':
  smemShape = (128, 32)
  # smemShape = (32, 128)
  elem_width = 2
  originSmem(smemShape, elem_width)
  # swizzleSmem(smemShape, elem_width, mode="our")
  # swizzleSmem(smemShape, elem_width, mode="tilelang")
  swizzleSmem(smemShape, elem_width, mode="our_other")
