from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import math
import random 

colors = ["BBFFFF", "AEEEEE", "96CDCD", "668B8B", "98F5FF", "8EE5EE", "7AC5CD", "53868B", "00F5FF", "00E5EE", "00C5CD", "00868B", "00FFFF", "00EEEE", "00CDCD", "008B8B", "97FFFF", "8DEEEE", "79CDCD", "528B8B", "7FFFD4", "76EEC6", "66CDAA", "458B74", "C1FFC1", "B4EEB4", "9BCD9B", "698B69", "54FF9F", "4EEE94", "43CD80", "2E8B57", "9AFF9A", "90EE90", "7CCD7C", "548B54", "00FF7F", "00EE76", "00CD66", "008B45", "00FF00", "00EE00", "00CD00", "008B00", "7FFF00", "76EE00", "66CD00", "458B00", "C0FF3E", "B3EE3A", "9ACD32", "698B22", "CAFF70", "BCEE68", "A2CD5A", "6E8B3D", "FFF68F", "EEE685", "CDC673", "8B864E", "FFEC8B", "EEDC82", "CDBE70", "8B814C", "FFFFE0", "EEEED1", "CDCDB4", "8B8B7A", "FFFF00", "EEEE00", "CDCD00", "8B8B00", "FFD700", "EEC900", "CDAD00", "8B7500", "FFC125", "EEB422", "CD9B1D", "8B6914", "FFB90F", "EEAD0E", "CD950C", "8B658B", "FFC1C1", "EEB4B4", "CD9B9B", "8B6969", "FF6A6A", "EE6363", "CD5555", "8B3A3A", "FF8247", "EE7942", "CD6839", "8B4726", "FFD39B", "EEC591", "CDAA7D", "8B7355", "FFE7BA", "EED8AE", "CDBA96", "8B7E66", "FFA54F", "EE9A49", "CD853F", "8B5A2B", "FF7F24", "EE7621", "CD661D", "8B4513", "FF3030", "EE2C2C", "CD2626", "8B1A1A", "FF4040", "EE3B3B", "CD3333", "8B2323", "FF8C69", "EE8262", "CD7054", "8B4C39", "FFA07A", "EE9572", "CD8162", "8B5742", "FFA500", "EE9A00", "CD8500", "8B5A00", "FF7F00", "EE7600", "CD6600", "8B4500", "FF7256", "EE6A50", "CD5B45", "8B3E2F", "FF6347", "EE5C42", "CD4F39", "8B3626", "FF4500", "EE4000", "CD3700", "8B2500", "FF0000", "EE0000", "CD0000", "8B0000", "FF1493", "EE1289", "CD1076", "8B0A50", "FF6EB4", "EE6AA7", "CD6090", "8B3A62", "FFB5C5", "EEA9B8", "CD919E", "8B636C", "FFAEB9", "EEA2AD", "CD8C95", "8B5F65", "FF82AB", "EE799F", "CD6889", "8B475D", "FF34B3", "EE30A7", "CD2990", "8B1C62", "FF3E96", "EE3A8C", "CD3278", "8B2252", "FF00FF", "EE00EE", "CD00CD", "8B008B", "FF83FA", "EE7AE9", "CD69C9", "8B4789", "FFBBFF", "EEAEEE", "CD96CD", "8B668B", "E066FF", "D15FEE", "B452CD", "7A378B", "BF3EFF", "B23AEE", "9A32CD", "68228B", "9B30FF", "912CEE", "7D26CD", "551A8B", "AB82FF", "9F79EE", "8968CD", "5D478B", "FFE1FF", "EED2EE", "CDB5CD", "8B7B8B", "1C1C1C", "363636", "4F4F4F", "696969", "828282", "9C9C9C", "B5B5B5", "CFCFCF", "E8E8E8", "A9A9A9", "00008B", "008B8B", "8B008B", "8B0000", "FFFAFA", "F8F8FF", "F5F5F5", "DCDCDC", "FFFAF0", "FDF5E6", "FAF0E6", "FAEBD7", "FFEFD5", "FFEBCD", "FFE4C4", "FFDAB9", "FFDEAD", "FFE4B5", "FFF8DC", "FFFFF0", "FFFACD", "FFF5EE", "F0FFF0", "F5FFFA", "F0FFFF", "F0F8FF", "E6E6FA", "FFF0F5", "FFE4E1", "FFFFFF", "000000", "2F4F4F", "696969", "708090", "778899", "BEBEBE", "D3D3D3", "191970", "000080", "6495ED", "483D8B", "6A5ACD", "7B68EE", "8470FF", "0000CD", "4169E1", "0000FF", "1E90FF", "00BFFF", "87CEEB", "87CEFA", "4682B4", "B0C4DE", "ADD8E6", "B0E0E6", "AFEEEE", "00CED1", "48D1CC", "40E0D0", "00FFFF", "E0FFFF", "5F9EA0", "66CDAA", "7FFFD4", "006400", "556B2F", "8FBC8F", "2E8B57", "3CB371", "20B2AA", "98FB98", "00FF7F", "7CFC00", "00FF00", "7FFF00", "00FA9A", "ADFF2F", "32CD32", "9ACD32", "228B22", "6B8E23", "BDB76B", "EEE8AA", "FAFAD2", "FFFFE0", "FFFF00", "FFD700", "EEDD82", "DAA520", "B8860B", "BC8F8F", "CD5C5C", "8B4513", "A0522D", "CD853F", "DEB887", "F5F5DC", "F5DEB3", "F4A460", "D2B48C", "D2691E", "B22222", "A52A2A", "E9967A", "FA8072", "FFA07A", "FFA500", "FF8C00", "FF7F50", "F08080", "FF6347", "FF4500", "FF0000", "FF69B4", "FF1493", "FFC0CB", "FFB6C1", "DB7093", "B03060", "C71585", "D02090", "FF00FF", "EE82EE", "DDA0DD", "DA70D6", "BA55D3", "9932CC", "9400D3", "8A2BE2", "A020F0", "9370DB", "D8BFD8", "FFFAFA", "EEE9E9", "CDC9C9", "8B8989", "FFF5EE", "EEE5DE", "CDC5BF", "8B8682", "FFEFDB", "EEDFCC", "CDC0B0", "8B8378", "FFE4C4", "EED5B7", "CDB79E", "8B7D6B", "FFDAB9", "EECBAD", "CDAF95", "8B7765", "FFDEAD", "EECFA1", "CDB38B", "8B795E", "FFFACD", "EEE9BF", "CDC9A5", "8B8970", "FFF8DC", "EEE8CD", "CDC8B1", "8B8878", "FFFFF0", "EEEEE0", "CDCDC1", "8B8B83", "F0FFF0", "E0EEE0", "C1CDC1", "838B83", "FFF0F5", "EEE0E5", "CDC1C5", "8B8386", "FFE4E1", "EED5D2", "CDB7B5", "8B7D7B", "F0FFFF", "E0EEEE", "C1CDCD", "838B8B", "836FFF", "7A67EE", "6959CD", "473C8B", "4876FF", "436EEE", "3A5FCD", "27408B", "0000FF", "0000EE", "0000CD", "00008B", "1E90FF", "1C86EE", "1874CD", "104E8B", "63B8FF", "5CACEE", "4F94CD", "36648B", "00BFFF", "00B2EE", "009ACD", "00688B", "87CEFF", "7EC0EE", "6CA6CD", "4A708B", "B0E2FF", "A4D3EE", "8DB6CD", "607B8B", "C6E2FF", "B9D3EE", "9FB6CD", "6C7B8B", "CAE1FF", "BCD2EE", "A2B5CD", "6E7B8B", "BFEFFF", "B2DFEE", "9AC0CD", "68838B", "E0FFFF", "D1EEEE"]
random.shuffle(colors)
# 只有相同颜色之间的才会产生bank conflict


def showMemory(table):
  wb = Workbook()
  sheet = wb.active
  thin_border = Side(border_style="thin", color="000000")
  cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
  cell_align = Alignment(horizontal="center")
  for i in range(len(table)):
    for j in range(len(table[i])):
      cell = sheet.cell(row=i+1, column=j+1, value=table[i][j][0])
      # print(colors[table[i][j][1]])
      cell.fill = PatternFill(start_color=colors[table[i][j][1]], fill_type="solid")
      cell.border = cell_border
      cell.alignment = cell_align
  wb.save("swizzle.xlsx")
  
def swizzle1(S, B, M, addr):
  # M：gran_width // elem_width 取2的对数，表示移动的位数
  # B：smem_shape[1] * elem_width // gran_width if < 8 else 8 取2的对数，表示移动位数
  # S：smem_shape[1] * elem_width // gran_width 取2的对数，表示移动位数
  # example：[16, 8] gran_withd
  # addr：xxxx|xxx|xxx
  #       2^4 |2^3|2^3
  #        S  | B | M 
  # S 取3位与B做异或替换B
  bmask = ((1 << B) - 1) << M
  return ((addr >> S) & bmask) ^ addr

def swizzle2(elem_width, gran_width, smem_shape, addr):
  gran_addr = addr * elem_width // gran_width
  gran_shape = (smem_shape[0], smem_shape[1] * elem_width // gran_width)  # (16, 2)
  gran_y = gran_addr // gran_shape[1]
  gran_x = gran_addr % gran_shape[1]
  num_cache_line_row = (128 // (smem_shape[1] * elem_width))  # 一个cache line 在shared row中占几行
  num_cache_line_col = (smem_shape[1] * elem_width // gran_width)  # shared memory 有多少列
  gran_x ^= (gran_y // num_cache_line_row) % num_cache_line_col  # 
  new_elem_addr = gran_y * smem_shape[1] + gran_x * gran_width // elem_width
  return new_elem_addr


def test():
  # shape:{16, 16} block_layout:{1, 1}
  table = [[(-1, -1) for _ in range(2)] for _ in range(16)]
  for tid in range(32):
    ty = tid % 16
    tx = tid // 16
    addr = ty * 16 + tx * 8
    # addr = swizzle2(2, 16, (16, 16), addr)
    M = math.log2(16 // 2)
    B = math.log2(16 // (16 // 2))
    S = math.log2((128 // (16 * 2)) * (16 // (16 // 2)))
    addr = swizzle1(int(S), int(B), int(M), addr)
    y = addr // 16
    x = (addr % 16) * 2 // 16
    table[y][x] = (tid, tid // 8)
  return table
    
    
def test2():
  # shape:{16, 32} block_layout:{1, 2}
  table = [[(-1, -1) for _ in range(4)] for _ in range(16)]
  for tid in range(64):
    warp_id = tid // 32
    lane_id = tid % 32
    
    ty = lane_id % 16
    tx = lane_id // 16
    addr = ty * 32 + (warp_id * 16) + tx * 8
    # addr = swizzle2(2, 16, (16, 32), addr)
    M = math.log2(16 // 2)
    B = math.log2(32 // (16 // 2))
    S = math.log2((128 // (32 * 2)) * (32 // (16 // 2)))
    addr = swizzle1(int(S), int(B), int(M), addr)
    y = addr // 32
    x = (addr % 32) * 2 // 16
    table[y][x] = (tid, tid // 8)
  return table


def test1():
  elem_width = 2
  gran_width = 16
  iteration_num = 2
  smem_shape = (32, 64)
  block_layout = (1, 4)
  warp_layout = (16, 2)
  thread_num = smem_shape[0] * smem_shape[1] * elem_width // gran_width // iteration_num
  # init table list
  table = [[(-1, -1) for _ in range(smem_shape[1] * elem_width // gran_width)] for _ in range(smem_shape[0])]
  for tid in range(thread_num):
    warp_id = tid // 32
    lane_id = tid % 32
    wy = warp_id // block_layout[1]
    wx = warp_id % block_layout[1]
    ly = lane_id % warp_layout[0]
    lx = lane_id // warp_layout[0]
    # block inner addr
    addr = (wy * warp_layout[0] + ly) * smem_shape[1] + wx * (warp_layout[1] * gran_width // elem_width) + lx * gran_width // elem_width
    # addr = swizzle2(elem_width, gran_width, smem_shape, addr)
    M = math.log2(gran_width // elem_width)
    B = math.log2(smem_shape[1] // (gran_width // elem_width))
    S = math.log2((128 // (smem_shape[1] * elem_width)) * (smem_shape[1] // (gran_width // elem_width)))
    addr = swizzle1(int(S), int(B), int(M), addr)
    for iter in range(iteration_num):
      offset = iter * (warp_layout[0] * smem_shape[1])
      addr += offset
      # show
      y = addr // smem_shape[1]
      x = (addr % smem_shape[1]) * elem_width // gran_width
      # print(y, x)
      table[y][x] = (tid, tid // 8)
  return table

def test3():
  # swizzle_128B deepgemm中的设置 表示两个atom中的排布
  table = [[(-1, -1) for _ in range(8)] for _ in range(16)]
  smem_shape = (16, 64)   # (16, 128) byte
  for tid in range(32):
    if tid < 16:
      y = tid % 16
      for i in range(8):  # in_atom_offset
        x = i
        x ^= y % 8
        # 虽然一个线程需要执行8次load
        # 但是这8次应该属于下一个周期，所以用其他颜色表达
        table[y][x] = (tid, i + (tid // 8) * 9)
  return table

def deepgemm_swizzle():
  # ldmatrix.2x.m8n8  swizzle mode: 128B
  table = [[(-1, -1) for _ in range(8)] for _ in range(512)]
  smem_shape = (512, 64)  # 物理 ，逻辑是(256, 128)
  for tid in range(256):  # 8*warp
    warp_id = tid // 32
    lane_id = tid % 32
    ty = lane_id
    if lane_id < 16:  # 2x的ldmatrix指令只有前16个线程有用
      for iter0 in range(16):   # 这迭代是每一个warp需要调用16次ldmatrix
        atom_offset = iter0 // 8   # 128B有8个16B
        in_atom_offset = iter0 % 8  # 每一个粒度列的索引
        tx = in_atom_offset
        warp_inner_addr = ty * smem_shape[1] + tx * 8   # 8个f16
        # warp_inner_addr = swizzle2(2, 16, smem_shape, warp_inner_addr)
        # M = math.log2(16 // 2)   # 粒度为16B，元素宽度2B
        # B = math.log2(smem_shape[1] // (16 // 2))
        # S = math.log2((128 // (smem_shape[1] * 2)) * (smem_shape[1] // (16 // 2)))
        warp_inner_addr = swizzle1(int(S), int(B), int(M), warp_inner_addr)
        for iter1 in range(2):  # 8个warp加载需要两波才能加载完成
          offset0 = atom_offset * 256 * smem_shape[1]  # 256是block_m
          offset1 = iter1 * 128 * smem_shape[1]  # 128是一波处理128行
          offset2 = warp_id * 16 * smem_shape[1]  # 16是一个warp处理的行数
          offset = offset0  + offset1 + offset2
          addr = offset + warp_inner_addr
          # show
          y = addr // smem_shape[1]
          x = (addr % smem_shape[1]) * 2 // 16  # 以16B为粒度画图
          table[y][x] = (tid, iter0 * 2 + iter1 + (tid // 8) * 9)
  return table
          
  
# 注：当这个shape的width超过128B时，需要进行reshape，不然SBM计算会报错
# 这个函数计算smb相当于在里面做了reshape
def getSBM(elem_width, gran_width, smem_shape):
  col_num = smem_shape[1] * elem_width // gran_width
  cache_line = col_num if col_num < 8 else 8  # cache line最长为128B，超过128B的长度不应该与这个部分进行swizzle操作
  S = math.log2(col_num)
  B = math.log2(cache_line)
  M = math.log(gran_width // 2)
  return S, B, M

if __name__ == '__main__':
  # table = test()  # 16*16 f16 1*warp
  # table = test2()  # 16*32 f16 2*warp
  # table = test1()  # 32*64 f16 4*warp
  # table = test3()  # 16*64 f16 1*warp
  table = deepgemm_swizzle()  # 512*64 8*warp
  showMemory(table=table)
    
      
